!fuser -k 5001/tcp >> /dev/null 2>&1
print("Initializing operational framework & engine environments...")
!pip -q install flask pandas numpy scikit-learn openpyxl

import time
import random
import io
import numpy as np
import pandas as pd
from flask import Flask, render_template_string, request, jsonify, Response, send_file
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from threading import Thread
from google.colab import output

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 2. SEED INTERNAL PROFILE POOL
# ==========================================
skills_pool = ["Python", "SQL", "ML", "Deep Learning", "NLP", "Data Analysis", "Pandas", "TensorFlow", "Statistics", "AWS", "PyTorch", "Docker", "Java", "C++"]
locations = ["Bangalore", "Hyderabad", "Chennai", "Mumbai", "Delhi", "Pune", "Kolkata"]
educations = ["Bachelor of Technology", "Bachelor of Engineering", "Master of Science", "Master of Technology"]

male_first_names = ["Rahul", "Amit", "Karan", "Vikram", "Rohan", "Siddharth", "Arjun", "Abhishek", "Deepak", "Sanjay"]
female_first_names = ["Priya", "Sneha", "Anjali", "Neha", "Divya", "Kriti", "Riya", "Pooja", "Tanvi", "Shreya"]
last_names = ["Kumar", "Sharma", "Verma", "Patil", "Singh", "Mehta", "Joshi", "Das", "Reddy", "Nair", "Mishra", "Gupta"]

candidates_pool = []
corpus_texts = []

top_5_presets = [
    {"name": "Rahul Kumar", "gender": "male", "exp": 6.2, "skills": ["Python", "ML", "SQL", "Deep Learning"], "seed": "jack"},
    {"name": "Priya Sharma", "gender": "female", "exp": 4.8, "skills": ["Python", "NLP", "ML", "TensorFlow"], "seed": "sara"},
    {"name": "Amit Verma", "gender": "male", "exp": 5.1, "skills": ["Python", "SQL", "ML", "Data Analysis"], "seed": "felix"},
    {"name": "Sneha Patil", "gender": "female", "exp": 3.9, "skills": ["Python", "ML", "NLP", "Pandas"], "seed": "zoey"},
    {"name": "Karan Singh", "gender": "male", "exp": 6.5, "skills": ["Python", "Deep Learning", "SQL"], "seed": "peter"}
]

tfidf_vectorizer = TfidfVectorizer(stop_words='english')
cand_tfidf_matrix = None

def generate_pool():
    global candidates_pool, corpus_texts, cand_tfidf_matrix
    candidates_pool = []
    corpus_texts = []
    
    random.seed(42)
    
    for i in range(140):
        if i < len(top_5_presets):
            name = top_5_presets[i]["name"]
            gender = top_5_presets[i]["gender"]
            exp = top_5_presets[i]["exp"]
            skills = top_5_presets[i]["skills"]
            seed_id = top_5_presets[i]["seed"]
        else:
            gender = "male" if random.random() > 0.5 else "female"
            first = random.choice(male_first_names) if gender == "male" else random.choice(female_first_names)
            name = f"{first} {random.choice(last_names)}"
            exp = round(random.uniform(1.5, 12.0), 1)
            skills = random.sample(skills_pool, random.randint(4, 7))
            seed_id = f"user_{i}_{random.randint(100,999)}"

        avatar_style = "avatar" if gender == "male" else "lorelei"
        avatar_url = f"https://api.dicebear.com/7.x/{avatar_style}/svg?seed={seed_id}&backgroundColor=0d1127,1e293b"

        candidate = {
            "candidate_id": f"CAND_{i+1:03d}",
            "name": name,
            "gender": gender,
            "location": random.choice(locations),
            "experience_years": exp,
            "skills": skills,
            "education": random.choice(educations),
            "summary": "Experienced AI and Machine Learning professional engineering complex production grade architectures.",
            "avatar": avatar_url
        }
        candidates_pool.append(candidate)
        corpus_texts.append(f"Skills: {' '.join(candidate['skills'])} Location: {candidate['location']} {candidate['summary']}")

    cand_tfidf_matrix = tfidf_vectorizer.fit_transform(corpus_texts)

generate_pool()

# ==========================================
# 3. FRONTEND UI TEMPLATE (HTML/CSS/JS)
# ==========================================
app = Flask(__name__)

INDEX_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>SMART AI Hiring System</title>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root {
            --bg-deep: #0b0f19;
            --bg-card: #111827;
            --bg-sidebar: #0f172a;
            --accent-blue: #3b82f6;
            --accent-green: #10b981;
            --accent-amber: #f59e0b;
            --text-main: #f3f4f6;
            --text-muted: #9ca3af;
            --border-color: #1f2937;
        }

        * { box-sizing: border-box; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }
        body { margin: 0; padding: 0; background-color: var(--bg-deep); color: var(--text-main); display: flex; height: 100vh; overflow: hidden; }

        .sidebar { width: 260px; background-color: var(--bg-sidebar); border-right: 1px solid var(--border-color); display: flex; flex-direction: column; padding: 20px 14px; justify-content: space-between; }
        .brand { display: flex; align-items: center; gap: 12px; padding: 10px; margin-bottom: 5px; }
        .brand i { font-size: 24px; color: var(--accent-blue); }
        .brand h2 { font-size: 15px; font-weight: 700; margin: 0; letter-spacing: 0.5px; }
        .sys-online { background-color: rgba(16, 185, 129, 0.1); color: var(--accent-green); font-size: 12px; font-weight: 600; padding: 4px 12px; border-radius: 20px; border: 1px solid rgba(16, 185, 129, 0.3); display: flex; align-items: center; gap: 6px; align-self: flex-start; margin-bottom: 25px; margin-left: 10px; }
        .sys-online::before { content: ''; width: 6px; height: 6px; background-color: var(--accent-green); border-radius: 50%; }

        .nav-links { display: flex; flex-direction: column; gap: 6px; flex-grow: 1; }
        .nav-item { display: flex; align-items: center; gap: 14px; padding: 12px 16px; color: #9ca3af; text-decoration: none; font-size: 14px; font-weight: 500; border-radius: 8px; transition: all 0.2s ease; cursor: pointer; }
        .nav-item:hover { background-color: #1f2937; color: var(--text-main); }
        .nav-item.active { background-color: var(--accent-blue); color: white; font-weight: 600; }

        .system-status-widget { background-color: rgba(17, 24, 39, 0.7); border: 1px solid var(--border-color); border-radius: 10px; padding: 14px; font-size: 13px; display: flex; flex-direction: column; gap: 10px; }
        .status-line { display: flex; justify-content: space-between; align-items: center; }
        .status-line span:first-child { color: var(--text-muted); }

        .workspace { flex: 1; display: flex; flex-direction: column; height: 100vh; overflow-y: auto; background-color: var(--bg-deep); }

        .kpi-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; padding: 24px 24px 12px 24px; }
        .kpi-card { background-color: var(--bg-card); border: 1px solid var(--border-color); border-radius: 12px; padding: 20px; display: flex; align-items: center; gap: 18px; }
        .kpi-icon { width: 48px; height: 48px; border-radius: 10px; display: flex; align-items: center; justify-content: center; font-size: 20px; }
        .kpi-icon.blue { background-color: rgba(59, 130, 246, 0.1); color: #3b82f6; }
        .kpi-icon.purple { background-color: rgba(168, 85, 247, 0.1); color: #a855f7; }
        .kpi-icon.green { background-color: rgba(16, 185, 129, 0.1); color: #10b981; }
        .kpi-icon.rose { background-color: rgba(244, 63, 94, 0.1); color: #f43f5e; }
        .kpi-info h4 { margin: 0; font-size: 11px; color: var(--text-muted); text-transform: uppercase; }
        .kpi-info h2 { margin: 4px 0 0 0; font-size: 26px; font-weight: 700; color: #ffffff; }

        .dashboard-grid { display: grid; grid-template-columns: 1.5fr 1.1fr; gap: 20px; padding: 12px 24px; flex-grow: 1; margin-bottom: 80px;}
        .left-col, .right-col { display: flex; flex-direction: column; gap: 20px; }

        .panel { background-color: var(--bg-card); border: 1px solid var(--border-color); border-radius: 12px; padding: 20px; display: flex; flex-direction: column; }
        .panel-header-row { display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px; }
        .panel-title { font-size: 13px; font-weight: 600; color: #9ca3af; margin: 0; text-transform: uppercase; letter-spacing: 0.5px; }

        .dropzone { border: 2px dashed #374151; border-radius: 10px; background-color: #1f2937; display: flex; flex-direction: column; align-items: center; justify-content: center; padding: 20px; text-align: center; height: 120px; cursor: pointer; }
        .dropzone p { font-size: 12px; color: var(--text-muted); margin: 0 0 10px 0; }
        
        .field-group { display: flex; flex-direction: column; gap: 6px; margin-bottom: 14px; }
        .field-group label { font-size: 12px; color: var(--text-muted); }
        .field-group input, .field-group textarea { background-color: #1f2937; border: 1px solid var(--border-color); border-radius: 8px; color: white; padding: 10px 14px; font-size: 13px; width: 100%; outline: none; }

        .btn-green { background-color: var(--accent-green); color: white; border: none; padding: 12px; border-radius: 8px; font-weight: 600; cursor: pointer; text-align: center; width: 100%; display: flex; align-items: center; justify-content: center; gap: 8px; transition: opacity 0.2s; }
        .btn-blue { background-color: var(--accent-blue); color: white; border: none; padding: 12px; border-radius: 8px; font-weight: 600; cursor: pointer; text-align: center; width: 100%; transition: opacity 0.2s; }
        .btn-amber { background-color: var(--accent-amber); color: #0b0f19; border: none; padding: 6px 12px; border-radius: 6px; font-size: 12px; font-weight: 700; cursor: pointer; display: flex; align-items: center; gap: 6px; text-decoration: none; transition: opacity 0.2s; }
        .btn-green:hover, .btn-blue:hover, .btn-amber:hover { opacity: 0.9; }

        .leaderboard-table-container { overflow-y: auto; background-color: #0f172a; border-radius: 10px; max-height: 380px; border: 1px solid var(--border-color); }
        .ld-table { width: 100%; border-collapse: collapse; text-align: left; font-size: 13px; }
        .ld-table th { background-color: #1e293b; padding: 12px; color: #9ca3af; position: sticky; top: 0; z-index: 10; }
        .ld-table td { padding: 10px 12px; border-bottom: 1px solid #1e2937; color: #e5e7eb; }

        .gender-badge { padding: 2px 8px; border-radius: 12px; font-size: 11px; font-weight: 600; text-transform: capitalize; }
        .gender-male { background-color: rgba(59, 130, 246, 0.15); color: #60a5fa; }
        .gender-female { background-color: rgba(244, 63, 94, 0.15); color: #f43f5e; }

        .top-match-item { display: flex; align-items: center; justify-content: space-between; padding: 12px; background-color: #1e293b; border-radius: 8px; margin-bottom: 8px; border-left: 3px solid var(--accent-blue); }
        .top-match-info { display: flex; align-items: center; gap: 10px; }
        .top-match-avatar { width: 28px; height: 28px; border-radius: 50%; background-color: #0d1127; border: 1px solid var(--border-color); }
        .score-val { color: var(--accent-green); font-weight: 700; }
        
        .chart-container { height: 180px; position: relative; }
    </style>
</head>
<body>

    <div class="sidebar">
        <div>
            <div class="brand">
                <i class="fa-solid fa-layer-group"></i>
                <h2>SMART AI HIRING SYSTEM</h2>
            </div>
            <div class="sys-online">System Online</div>

            <div class="nav-links">
                <a class="nav-item active" id="nav-dash" onclick="switchTab('dash')"><i class="fa-solid fa-chart-pie"></i> Dashboard</a>
                <a class="nav-item" id="nav-upload" onclick="switchTab('upload')"><i class="fa-solid fa-cloud-arrow-up"></i> Upload Resumes</a>
                <a class="nav-item" id="nav-rank" onclick="switchTab('rank')"><i class="fa-solid fa-arrow-down-short-wide"></i> Rank Candidates</a>
                <a class="nav-item" id="nav-lead" onclick="switchTab('lead')"><i class="fa-solid fa-crown"></i> Leaderboard</a>
            </div>
        </div>

        <div class="system-status-widget">
            <div class="status-line"><span>Server Status:</span><span style="color:var(--accent-green)">Running</span></div>
            <div class="status-line"><span>Total Pool:</span><span id="stat-pool-size" style="color:white; font-weight:bold;">140</span></div>
            <div class="status-line"><span>Target Scope:</span><span style="color:white; font-weight:bold;">Active alignment</span></div>
        </div>
    </div>

    <div class="workspace">
        <div class="kpi-row">
            <div class="kpi-card">
                <div class="kpi-icon blue"><i class="fa-solid fa-users"></i></div>
                <div class="kpi-info"><h4>Total Pool Profiles</h4><h2 id="kpi-pool-size">140</h2></div>
            </div>
            <div class="kpi-card">
                <div class="kpi-icon purple"><i class="fa-solid fa-business-time"></i></div>
                <div class="kpi-info"><h4>Avg Experience</h4><h2>6.4 Years</h2></div>
            </div>
            <div class="kpi-card">
                <div class="kpi-icon green"><i class="fa-solid fa-circle-check"></i></div>
                <div class="kpi-info"><h4>Selected Metrics</h4><h2 id="kpi-ranked-count">140 Ranked</h2></div>
            </div>
            <div class="kpi-card">
                <div class="kpi-icon rose"><i class="fa-solid fa-star"></i></div>
                <div class="kpi-info"><h4>Top Score Bound</h4><h2>98.4%</h2></div>
            </div>
        </div>

        <div class="dashboard-grid">
            <div class="left-col">
                <div class="panel" id="panel-upload-view">
                    <div class="panel-header-row"><div class="panel-title">Upload Real Resume (.txt File)</div></div>
                    <div class="dropzone" onclick="document.getElementById('hidden-file-input').click()">
                        <p>Drag & Drop candidate .txt resume file here or</p>
                        <button type="button" style="background:#3b82f6; border:none; padding:6px 12px; color:white; border-radius:4px; cursor:pointer;">Choose File</button>
                        <input type="file" id="hidden-file-input" style="display:none;" accept=".txt" onchange="handleFileSelected(this)">
                        <div id="file-lbl" style="margin-top:10px; color:#f59e0b; font-size:12px; font-weight:bold;">No file selected</div>
                    </div>
                    <button class="btn-green" style="margin-top:15px;" onclick="uploadFileToServer()">Process & Ingest Resume</button>
                </div>

                <div class="panel" id="panel-jd-view">
                    <div class="panel-header-row"><div class="panel-title">Job Description Engine</div></div>
                    <div class="field-group">
                        <label>Target Role Specifications</label>
                        <textarea id="jd-text" rows="4">Looking for an AI engineer capable of deploying real-time analysis tools, working with Python, ML, SQL, and Deep Learning infrastructures.</textarea>
                    </div>
                    <button class="btn-blue" onclick="runMatching()">Execute Rank Alignment</button>
                </div>

                <div class="panel">
                    <div class="panel-header-row"><div class="panel-title">Skill Distribution (Top Core Fields)</div></div>
                    <div class="chart-container">
                        <canvas id="skillsChart"></canvas>
                    </div>
                </div>
            </div>

            <div class="right-col">
                <div class="panel">
                    <div class="panel-header-row"><div class="panel-title">Top Position Matches</div></div>
                    <div id="top-matches-stack"></div>
                </div>

                <div class="panel" id="panel-lead-view">
                    <div class="panel-header-row">
                        <div class="panel-title">Leaderboard Matrix</div>
                        <button class="btn-amber" onclick="downloadExcel()"><i class="fa-solid fa-file-excel"></i> Export Excel</button>
                    </div>
                    <div class="leaderboard-table-container">
                        <table class="ld-table">
                            <thead>
                                <tr><th>ID</th><th>Name</th><th>Gender</th><th>Score</th></tr>
                            </thead>
                            <tbody id="lead-table-body"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script>
        let globalSelectedFile = null;

        function handleFileSelected(input) {
            if(input.files && input.files[0]) {
                globalSelectedFile = input.files[0];
                document.getElementById('file-lbl').innerText = "Selected: " + globalSelectedFile.name;
            }
        }

        function uploadFileToServer() {
            if(!globalSelectedFile) {
                alert("Please pick a valid resume .txt file first!");
                return;
            }
            let formData = new FormData();
            formData.append("resume", globalSelectedFile);

            fetch('/upload_resume_api', {
                method: 'POST',
                body: formData
            })
            .then(res => res.json())
            .then(data => {
                if(data.success) {
                    alert("Real-time parsing complete! Added as candidate: " + data.candidate_id);
                    document.getElementById('stat-pool-size').innerText = data.total_count;
                    document.getElementById('kpi-pool-size').innerText = data.total_count;
                    document.getElementById('kpi-ranked-count').innerText = data.total_count + " Ranked";
                    document.getElementById('file-lbl').innerText = "No file selected";
                    globalSelectedFile = null;
                    runMatching();
                } else {
                    alert("Error parsing file: " + data.message);
                }
            })
            .catch(err => alert("Upload failed: " + err));
        }

        function switchTab(tab) {
            document.querySelectorAll('.nav-item').forEach(el => el.classList.remove('active'));
            if(tab === 'dash') document.getElementById('nav-dash').classList.add('active');
            if(tab === 'upload') document.getElementById('nav-upload').classList.add('active');
            if(tab === 'rank') document.getElementById('nav-rank').classList.add('active');
            if(tab === 'lead') document.getElementById('nav-lead').classList.add('active');
        }

        function runMatching() {
            fetch('/match_api', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({desc: document.getElementById('jd-text').value})
            })
            .then(res => res.json())
            .then(data => {
                let matchesHtml = "";
                let tableHtml = "";
                
                data.slice(0, 5).forEach(c => {
                    matchesHtml += `
                    <div class="top-match-item">
                        <div class="top-match-info">
                            <img class="top-match-avatar" src="${c.avatar}" alt="avatar">
                            <span><b>${c.name}</b> (${c.exp} Yrs Exp)</span>
                        </div>
                        <span class="score-val">${c.score}%</span>
                    </div>`;
                });
                document.getElementById('top-matches-stack').innerHTML = matchesHtml;

                data.forEach(c => {
                    let gBadge = c.gender === 'male' ? '<span class="gender-badge gender-male">Male</span>' : '<span class="gender-badge gender-female">Female</span>';
                    tableHtml += `<tr><td>${c.id}</td><td>${c.name}</td><td>${gBadge}</td><td style="color:#10b981; font-weight:bold;">${c.score}%</td></tr>`;
                });
                document.getElementById('lead-table-body').innerHTML = tableHtml;
            });
        }

        function downloadExcel() {
            const jdText = encodeURIComponent(document.getElementById('jd-text').value);
            window.location.href = `/download_excel?desc=${jdText}`;
        }

        window.onload = function() {
            const ctx = document.getElementById('skillsChart').getContext('2d');
            new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: ['Python', 'SQL', 'ML', 'Deep Learning', 'NLP', 'Pandas', 'AWS'],
                    datasets: [{
                        data: [94, 87, 81, 74, 65, 60, 55],
                        backgroundColor: '#3b82f6',
                        borderRadius: 4
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: { legend: { display: false } },
                    scales: {
                        y: { grid: { color: '#374151' }, ticks: { color: '#9ca3af' } },
                        x: { grid: { display: false }, ticks: { color: '#9ca3af' } }
                    }
                }
            });
            runMatching();
        };
    </script>
</body>
</html>
"""

# ==========================================
# 4. SERVER ROUTING API & EXCEL EXPORT
# ==========================================
@app.route('/')
def home():
    return render_template_string(INDEX_TEMPLATE)

@app.route('/upload_resume_api', methods=['POST'])
def upload_resume_api():
    global candidates_pool, corpus_texts, cand_tfidf_matrix
    if 'resume' not in request.files:
        return jsonify({"success": False, "message": "No file chunk found"})
    
    file = request.files['resume']
    if file.filename == '':
        return jsonify({"success": False, "message": "Empty filename selection"})
    
    try:
        # Read file text data in real time
        text_content = file.read().decode('utf-8', errors='ignore')
        
        # Build out structured identity profile matching the vector corpus layout
        next_id = len(candidates_pool) + 1
        cand_id_str = f"CAND_{next_id:03d}"
        
        # Simple extraction heuristics from provided content body
        inferred_name = file.filename.rsplit('.', 1)[0].replace('_', ' ').replace('-', ' ').title()
        
        new_candidate = {
            "candidate_id": cand_id_str,
            "name": inferred_name,
            "gender": "female" if any(x in text_content.lower() for x in ["she", "her", "sara", "priya"]) else "male",
            "location": random.choice(locations),
            "experience_years": round(random.uniform(3.0, 8.5), 1),
            "skills": [s for s in skills_pool if s.lower() in text_content.lower()] or ["Python", "Data Science"],
            "education": random.choice(educations),
            "summary": text_content[:150].strip().replace('\n', ' ') + "...",
            "avatar": f"https://api.dicebear.com/7.x/avatar/svg?seed={inferred_name}&backgroundColor=0d1127,1e293b"
        }
        
        # Update our active model corpus pools instantly
        candidates_pool.append(new_candidate)
        corpus_texts.append(f"Skills: {' '.join(new_candidate['skills'])} Location: {new_candidate['location']} RawResume: {text_content}")
        
        # Re-train and re-index vector spaces on the fly
        cand_tfidf_matrix = tfidf_vectorizer.fit_transform(corpus_texts)
        
        return jsonify({
            "success": True, 
            "candidate_id": cand_id_str, 
            "total_count": len(candidates_pool)
        })
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

def compute_scores(desc):
    global cand_tfidf_matrix
    query_tfidf = tfidf_vectorizer.transform([desc])
    similarities = cosine_similarity(query_tfidf, cand_tfidf_matrix).flatten()
    
    scored = []
    for idx, cand in enumerate(candidates_pool):
        base_score = similarities[idx] * 100
        final_score = round(45 + (base_score * 0.45) + (cand['experience_years'] * 0.8), 1)
        final_score = min(final_score, 98.4)
        
        cand_data = cand.copy()
        cand_data['score'] = final_score
        scored.append(cand_data)
        
    scored.sort(key=lambda x: x['score'], reverse=True)
    return scored

@app.route('/match_api', methods=['POST'])
def match_api():
    req = request.get_json() or {}
    desc = req.get('desc', '')
    scored = compute_scores(desc)
    
    frontend_data = [{
        "id": c["candidate_id"],
        "name": c["name"],
        "gender": c["gender"],
        "exp": c["experience_years"],
        "score": c["score"],
        "avatar": c["avatar"]
    } for c in scored]
    
    return jsonify(frontend_data)

@app.route('/download_excel', methods=['GET'])
def download_excel():
    desc = request.args.get('desc', '')
    scored_candidates = compute_scores(desc)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked AI Candidates"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    male_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    female_fill = PatternFill(start_color="FDF2F8", end_color="FDF2F8", fill_type="solid")

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="111827")
    bold_font = Font(name="Segoe UI", size=10, bold=True, color="111827")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    headers = [
        "Candidate ID", "Candidate Name", "Gender", "Location", 
        "Experience (Yrs)", "Education Level", "Key Technical Skills", "Avatar URL Profile Pic", "Match Alignment"
    ]

    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for r_idx, cand in enumerate(scored_candidates, 2):
        ws.append([
            cand["candidate_id"],
            cand["name"],
            cand["gender"].capitalize(),
            cand["location"],
            cand["experience_years"],
            cand["education"],
            ", ".join(cand["skills"]),
            cand["avatar"],
            cand["score"] / 100.0
        ])
        
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            
            if c_idx == 3:
                cell.fill = male_fill if cand["gender"] == "male" else female_fill
                cell.alignment = Alignment(horizontal="center")
            elif r_idx % 2 == 0:
                cell.fill = zebra_fill
                
            if c_idx in [1, 4, 5]:
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 9:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '0.0%'
                cell.font = bold_font
            elif c_idx == 8:
                cell.font = Font(name="Segoe UI", size=9, color="2563EB", underline="single")

        ws.row_dimensions[r_idx].height = 20

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    ws.column_dimensions['G'].width = 38
    ws.column_dimensions['H'].width = 25

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        as_attachment=True,
        download_name="Smart_AI_Ranked_Candidates.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==========================================
# 5. EXECUTE APP RUNTIME
# ==========================================
def run_app():
    app.run(host='0.0.0.0', port=5001, threaded=True)

server_thread = Thread(target=run_app)
server_thread.daemon = True
server_thread.start()

time.sleep(2)
colab_env_domain = output.eval_js("google.colab.kernel.proxyPort(5001)")

print("\n" + "="*70)
print(f"🚀 DASHBOARD RECONFIGURED SUCCESSFULLY!")
print(f"👉 CLICK LINK BELOW TO INTERACT FULL SCREEN:")
print(f"   {colab_env_domain}")
print("="*70 + "\n")

output.serve_kernel_port_as_iframe(5001, height=850)
